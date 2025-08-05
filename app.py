import json
import pandas as pd
import streamlit as st
import io
import os
from datetime import datetime

# --- Excel 格式化函式庫 ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ==============================================================================
# --- 核心功能函式 (來自前一版，已整合) ---
# ==============================================================================

def extract_json_from_text(raw_text):
    """從包含額外文字的字串中，智慧地找出並擷取JSON內容。"""
    first_bracket = -1
    first_curly = raw_text.find('{')
    first_square = raw_text.find('[')
    if first_curly == -1 and first_square == -1: raise ValueError("在檔案中找不到有效的JSON起始符號 '{' 或 '['。")
    if first_curly != -1 and first_square != -1: first_bracket = min(first_curly, first_square)
    elif first_curly != -1: first_bracket = first_curly
    else: first_bracket = first_square
    last_bracket = -1
    last_curly = raw_text.rfind('}')
    last_square = raw_text.rfind(']')
    if last_curly == -1 and last_square == -1: raise ValueError("在檔案中找不到有效的JSON結束符號 '}' 或 ']'。")
    if last_curly != -1 and last_square != -1: last_bracket = max(last_curly, last_square)
    elif last_curly != -1: last_bracket = last_curly
    else: last_bracket = last_square
    if first_bracket > last_bracket: raise ValueError("JSON的起始與結束符號位置不正確。")
    json_string = raw_text[first_bracket : last_bracket + 1]
    return json_string

def create_personal_summary_sheet(wb, data, font_name, fallback_font, font_size):
    """創建個人申請資料專用的摘要工作表。"""
    ws = wb.create_sheet("資料摘要")
    summary_data = [
        ["項目", "數值"],
        ["總申請人數", len(data)],
        ["男性", len([d for d in data if d.get("性別") == "男"])],
        ["女性", len([d for d in data if d.get("性別") == "女"])],
        ["同意", len([d for d in data if d.get("勞動部檢核結果") == "同意"])],
        ["承辦中", len([d for d in data if d.get("勞動部檢核結果") == "承辦中"])],
        ["軟體技術開發", len([d for d in data if d.get("子領域") == "軟體技術開發"])],
        ["數位科技內容產製或擴散", len([d for d in data if d.get("子領域") == "數位科技內容產製或擴散"])],
    ]
    header_font = Font(bold=True, color="FFFFFF", name=font_name, size=font_size)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell_font = Font(name=font_name, size=font_size)
    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font; cell.fill = header_fill
            else:
                cell.font = cell_font
    try:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font.name != font_name:
                    cell.font = Font(name=fallback_font, size=font_size, bold=cell.font.bold, color=cell.font.color)
    except:
        pass
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

def generate_formatted_excel_bytes(json_data, include_summary=False):
    """
    核心的轉換與格式化引擎，接收 JSON 資料，返回 Excel 檔案的 bytes。
    Args:
        json_data (list or dict): 從JSON解析後的Python資料。
        include_summary (bool): 是否要加入個人申請資料的摘要頁。
    """
    if isinstance(json_data, dict):
        json_data = [json_data]
    
    df = pd.DataFrame(json_data)
    wb = Workbook()
    ws = wb.active
    ws.title = "資料內容"
    
    font_name = "Microsoft JhengHei"
    fallback_font = "Arial"
    font_size = 14
    header_font = Font(bold=True, color="FFFFFF", name=font_name, size=font_size)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name=font_name, size=font_size)
    cell_align = Alignment(vertical="top", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align
            else:
                cell.font = cell_font; cell.alignment = cell_align
            cell.border = border
    
    try:
        for row in ws.iter_rows():
            for cell in row:
                if cell.font.name != font_name:
                    cell.font = Font(name=fallback_font, size=font_size, bold=cell.font.bold, color=cell.font.color)
    except:
        pass
    
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                length = sum(2 if ord(char) > 127 else 1 for char in str(cell.value))
                if length > max_length: max_length = length
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 8), 60)

    for row in ws.iter_rows(min_row=2):
        lines = max(str(cell.value).count('\n') + 1 for cell in row if cell.value) if any(cell.value for cell in row) else 1
        ws.row_dimensions[row[0].row].height = min(max(lines * 15, 15), 200)

    ws.freeze_panes = "A2"

    if include_summary:
        try:
            create_personal_summary_sheet(wb, json_data, font_name, fallback_font, font_size)
        except Exception as summary_error:
            st.warning(f"⚠️ 無法生成摘要頁，可能缺少必要欄位。錯誤：{summary_error}")

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    return output_buffer.getvalue()

# ==============================================================================
# --- Streamlit App 主流程 (已升級) ---
# ==============================================================================
def main():
    st.set_page_config(page_title="智慧型轉檔工具", page_icon="🔄")
    st.title("🔄 智慧型 JSON to Excel 轉換工具")
    st.write("本工具能自動解析含有額外文字的檔案，並提供通用或專用的Excel格式化選項。")

    # --- 修改：更新選項以符合雙軌制 ---
    conversion_type = st.radio(
        "請選擇符合您資料類型的功能：",
        ('通用格式化轉檔 (適用任何資料)', '個人申請資料分析 (含摘要頁)'),
        horizontal=True
    )

    uploaded_file = st.file_uploader(
        "請上傳您的 JSON 或 TXT 檔案",
        type=['json', 'txt']
    )

    if uploaded_file is not None:
        try:
            # --- 修改：加入智慧擷取功能 ---
            # 1. 將上傳的檔案內容讀取為字串
            raw_content = uploaded_file.getvalue().decode("utf-8")
            # 2. 智慧擷取純JSON部分
            json_string = extract_json_from_text(raw_content)
            # 3. 從純JSON字串解析成Python物件
            json_data = json.loads(json_string)

            st.success("✅ 檔案解析成功！請點擊下方按鈕下載轉換結果。")

            excel_bytes = None
            file_suffix = ""
            
            # --- 修改：根據新的選項呼叫核心引擎 ---
            if conversion_type == '通用格式化轉檔 (適用任何資料)':
                excel_bytes = generate_formatted_excel_bytes(json_data, include_summary=False)
                file_suffix = "formatted"

            elif conversion_type == '個人申請資料分析 (含摘要頁)':
                excel_bytes = generate_formatted_excel_bytes(json_data, include_summary=True)
                file_suffix = "personal_analysis"

            if excel_bytes:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                original_filename = os.path.splitext(uploaded_file.name)[0]
                
                st.download_button(
                    label=f"📥 點此下載結果",
                    data=excel_bytes,
                    file_name=f"{original_filename}_{file_suffix}_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"處理檔案時發生錯誤：{e}")
            st.warning("請確認檔案內容是否包含有效的 JSON 格式，或嘗試聯絡管理員。")

if __name__ == "__main__":
    main()
